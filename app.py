from flask import Flask, request, send_file
import openpyxl
import tempfile

app = Flask(__name__)

@app.route("/fill-excel", methods=["POST"])
def fill_excel():
    data = request.json

    # Load your template
    wb = openpyxl.load_workbook("template.xlsx")
    ws = wb["Template"]  # sheet name in your Excel file

    # Fill specific cells (preserving formulas/dropdowns)
    ws["B3"] = data.get("pay_period", "September - October")
    ws["B4"] = data.get("employee_name", "")
    ws["A13"] = data.get("project", "Woodleaze")
    ws["B13"] = data.get("shift_type", "Day Shift")
    ws["C13"] = data.get("start_time", "22/09/2024 08:00")
    ws["D13"] = data.get("end_time", "22/09/2024 17:00")
    ws["G13"] = data.get("pay_rate", 1)
    ws["B77"] = data.get("signature", "Ayoor")

    # Save to a temporary file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return send_file(tmp.name, as_attachment=True, download_name="filled_form.xlsx")

@app.route("/", methods=["GET"])
def home():
    return {"message": "✅ Excel Backend is running!"}

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
